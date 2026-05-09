#!/usr/bin/env python3
"""
顧客提供 障害票.xlsx テンプレートに値を流し込んで `output/<INC-ID>/障害票.xlsx` を生成。

Usage:
    python scripts/fill_template.py \
        --template templates/障害票_template.xlsx \
        --output   output/INC-202604-001/障害票.xlsx \
        --json     '{"incident_id":"...","title":"...","phenomenon":"..."}'

または stdin 経由:
    cat fields.json | python scripts/fill_template.py \
        --template templates/障害票_template.xlsx \
        --output   output/INC-202604-001/障害票.xlsx

CELL_MAP は顧客テンプレに合わせて調整してください。
named_range が定義されているテンプレなら NAMED_RANGE 側を使う方が堅牢。

依存:
    pip install openpyxl
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

# ───────────────────────────────────────────────────────────────────────────
# テンプレ毎に調整するセクション
# ───────────────────────────────────────────────────────────────────────────
# 顧客テンプレが固定セル番地式の場合 (sheet 名: "障害票", セル: "C3" 等)
CELL_MAP: dict[str, tuple[str, str]] = {
    "incident_id":   ("障害票", "C3"),
    "title":         ("障害票", "C4"),
    "reported_by":   ("障害票", "C5"),
    "reported_at":   ("障害票", "F5"),
    "severity":      ("障害票", "C6"),
    "phenomenon":    ("障害票", "C8"),
    "repro_steps":   ("障害票", "C12"),
    "design_refs":   ("障害票", "C18"),
    "code_refs":     ("障害票", "C24"),
    "root_cause":    ("障害票", "C30"),
    "design_drift":  ("障害票", "C36"),
    "impact":        ("障害票", "C40"),
    "countermeasure_a": ("障害票", "C46"),
    "countermeasure_b": ("障害票", "C52"),
    "recommendation": ("障害票", "C58"),
    "review_notes":  ("障害票", "C62"),
}

# 顧客テンプレが named_range 式の場合（推奨、堅牢）
# 使う場合は CELL_MAP を空にして NAMED_RANGE_MAP を埋める
NAMED_RANGE_MAP: dict[str, str] = {
    # "incident_id": "INCIDENT_ID",
    # "title":       "TITLE",
    # ...
}


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="障害票テンプレに値を流し込む")
    p.add_argument("--template", type=Path, required=True)
    p.add_argument("--output", type=Path, required=True)
    p.add_argument("--json", type=str,
                   help="フィールド JSON 文字列。未指定時は stdin から読込")
    return p.parse_args()


def load_fields(args: argparse.Namespace) -> dict:
    if args.json:
        return json.loads(args.json)
    raw = sys.stdin.read()
    if not raw.strip():
        log.error("--json も stdin もありません")
        sys.exit(1)
    return json.loads(raw)


def fill_by_cell(wb, fields: dict) -> int:
    """CELL_MAP に従い固定セルに書き込む."""
    written = 0
    for key, (sheet_name, cell_addr) in CELL_MAP.items():
        if key not in fields:
            continue
        if sheet_name not in wb.sheetnames:
            log.warning("シート '%s' が存在しません、skip", sheet_name)
            continue
        ws = wb[sheet_name]
        ws[cell_addr] = str(fields[key])
        written += 1
    return written


def fill_by_named_range(wb, fields: dict) -> int:
    """NAMED_RANGE_MAP に従い named range に書き込む."""
    written = 0
    for key, range_name in NAMED_RANGE_MAP.items():
        if key not in fields:
            continue
        defined = wb.defined_names.get(range_name)
        if defined is None:
            log.warning("named range '%s' が存在しません、skip", range_name)
            continue
        for sheet_name, cell_range in defined.destinations:
            ws = wb[sheet_name]
            for row in ws[cell_range]:
                for cell in row:
                    cell.value = str(fields[key])
                    written += 1
                    break
                break
    return written


def main() -> int:
    try:
        from openpyxl import load_workbook
    except ImportError:
        log.error("openpyxl が未インストール: `pip install openpyxl`")
        return 1

    args = parse_args()
    if not args.template.exists():
        log.error("テンプレ存在せず: %s", args.template)
        return 1

    fields = load_fields(args)
    log.info("入力フィールド: %d 件", len(fields))

    keep_vba = args.template.suffix.lower() == ".xlsm"
    wb = load_workbook(args.template, keep_vba=keep_vba)

    written = 0
    if NAMED_RANGE_MAP:
        written += fill_by_named_range(wb, fields)
    written += fill_by_cell(wb, fields)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    log.info("出力: %s (%d セル書込)", args.output, written)
    return 0


if __name__ == "__main__":
    sys.exit(main())
